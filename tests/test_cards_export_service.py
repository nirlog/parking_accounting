from __future__ import annotations

from datetime import date, datetime
from importlib.util import find_spec
from pathlib import Path
import tempfile
import unittest

from parking_app.services.cards_export_service import (
    build_cards_export_rows,
    cards_export_columns,
)
from parking_app.services.cards_table_service import CardTableRow
from parking_app.services.export_service import export_rows_to_xlsx

OPENPYXL_AVAILABLE = find_spec("openpyxl") is not None
if OPENPYXL_AVAILABLE:
    from openpyxl import load_workbook


class CardsExportServiceTests(unittest.TestCase):
    def _row(self, **kw) -> CardTableRow:
        data = dict(
            card_id=1,
            card_number="000001",
            paper_card_number="147",
            place_number="101",
            fio="Иванов Иван Иванович",
            state_number="А123АА178",
            vehicle="Lada Vesta",
            phone="79211112233",
            paid_until=date(2026, 5, 31),
            payment_status="Оплачено",
            card_status="active",
        )
        data.update(kw)
        return CardTableRow(**data)

    def test_build_cards_export_rows_formats_dates_and_status(self):
        out = build_cards_export_rows([self._row()])[0]
        self.assertEqual(out["paid_until"], "31.05.2026")
        self.assertEqual(out["card_status"], "Активная")

    def test_build_cards_export_rows_handles_no_payments(self):
        out = build_cards_export_rows([self._row(paid_until=None)])[0]
        self.assertEqual(out["paid_until"], "Нет оплат")

    def test_build_cards_export_rows_formats_closed_and_archived_status(self):
        closed = build_cards_export_rows([self._row(card_status="closed")])[0]
        archived = build_cards_export_rows([self._row(card_status="archived")])[0]
        self.assertEqual(closed["card_status"], "Закрыта")
        self.assertEqual(archived["card_status"], "Архив")

    def test_cards_export_columns_order(self):
        keys = [c.key for c in cards_export_columns()]
        self.assertEqual(keys, [
            "card_number", "paper_card_number", "place_number", "fio", "state_number", "vehicle",
            "phone", "paid_until", "payment_status", "card_status"
        ])

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
    def test_export_service_sanitizes_formula_values_for_cards(self):
        row = self._row(vehicle="=1+1")
        export_rows = build_cards_export_rows([row])
        with tempfile.TemporaryDirectory() as td:
            out = export_rows_to_xlsx(
                output_dir=Path(td),
                report_name="cards",
                sheet_name="Карточки",
                columns=cards_export_columns(),
                rows=export_rows,
                now=datetime(2026, 5, 21, 12, 0),
            )
            ws = load_workbook(out)["Карточки"]
            self.assertEqual(ws["F2"].value, "'=1+1")


if __name__ == "__main__":
    unittest.main()
