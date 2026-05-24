from __future__ import annotations

from datetime import date, datetime
from importlib.util import find_spec
from pathlib import Path
import tempfile
import unittest

from parking_app.services.export_service import export_rows_to_xlsx
from parking_app.services.places_export_service import build_places_export_rows, places_export_columns
from parking_app.services.places_table_service import PlaceTableRow

OPENPYXL_AVAILABLE = find_spec("openpyxl") is not None
if OPENPYXL_AVAILABLE:
    from openpyxl import load_workbook


class PlacesExportServiceTests(unittest.TestCase):
    def _row(self, *, display_status: str = "free", paid_until: date | None = None, client_fio: str = "—", note: str = "—") -> PlaceTableRow:
        return PlaceTableRow(
            place_id=1,
            place_number="101",
            base_status="free",
            display_status=display_status,
            client_fio=client_fio,
            state_number="А123АА178" if display_status == "occupied" else "—",
            vehicle="Toyota Camry" if display_status == "occupied" else "—",
            paid_until=paid_until,
            payment_status="Оплачено" if paid_until else ("Нет оплат" if display_status == "occupied" else "—"),
            note=note,
        )

    def test_build_places_export_rows_formats_statuses(self) -> None:
        rows = [
            self._row(display_status="free"),
            self._row(display_status="occupied", paid_until=date(2026, 5, 31)),
            self._row(display_status="reserved"),
            self._row(display_status="repair"),
        ]
        out = build_places_export_rows(rows)
        self.assertEqual([r["display_status"] for r in out], ["Свободно", "Занято", "Бронь", "Ремонт"])

    def test_build_places_export_rows_formats_paid_until(self) -> None:
        rows = [
            self._row(display_status="occupied", paid_until=date(2026, 5, 31)),
            self._row(display_status="occupied", paid_until=None),
            self._row(display_status="free", paid_until=None),
        ]
        out = build_places_export_rows(rows)
        self.assertEqual(out[0]["paid_until"], "31.05.2026")
        self.assertEqual(out[1]["paid_until"], "Нет оплат")
        self.assertEqual(out[2]["paid_until"], "—")

    def test_places_export_columns_order(self) -> None:
        keys = [c.key for c in places_export_columns()]
        self.assertEqual(
            keys,
            [
                "place_number",
                "display_status",
                "client_fio",
                "state_number",
                "vehicle",
                "paid_until",
                "payment_status",
                "note",
            ],
        )

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
    def test_export_service_sanitizes_formula_values_for_places(self) -> None:
        export_rows = build_places_export_rows([self._row(display_status="occupied", paid_until=None, client_fio="=1+1")])
        with tempfile.TemporaryDirectory() as td:
            out = export_rows_to_xlsx(
                output_dir=Path(td),
                report_name="places",
                sheet_name="Места",
                columns=places_export_columns(),
                rows=export_rows,
                now=datetime(2026, 5, 21, 8, 30),
            )
            wb = load_workbook(out)
            ws = wb["Места"]
            self.assertEqual(ws["C2"].value, "'=1+1")


if __name__ == "__main__":
    unittest.main()
