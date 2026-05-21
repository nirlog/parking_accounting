from __future__ import annotations

from importlib.util import find_spec
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

OPENPYXL_AVAILABLE = find_spec("openpyxl") is not None

if OPENPYXL_AVAILABLE:
    from openpyxl import load_workbook

from parking_app.services.export_service import (
    ExportColumn,
    export_rows_to_xlsx,
    format_amount_rub,
    format_date_ddmmyyyy,
    make_export_filename,
    ensure_unique_export_path,
)


class ExportServiceTests(unittest.TestCase):
    def test_filename_format(self) -> None:
        name = make_export_filename("payments", datetime(2026, 5, 21, 8, 30))
        self.assertEqual(name, "payments_2026-05-21_0830.xlsx")

    def test_format_date(self) -> None:
        self.assertEqual(format_date_ddmmyyyy(date(2026, 5, 21)), "21.05.2026")
        self.assertEqual(format_date_ddmmyyyy(None), "")

    def test_format_amount(self) -> None:
        self.assertEqual(format_amount_rub(800000), "8000.00")
        self.assertEqual(format_amount_rub(None), "")


    def test_ensure_unique_export_path(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            base = root / "payments_2026-05-21_0830.xlsx"
            base.write_bytes(b"a")
            unique = ensure_unique_export_path(base)
            self.assertEqual(unique.name, "payments_2026-05-21_0830_1.xlsx")

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
    def test_export_with_empty_rows_writes_marker(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = export_rows_to_xlsx(
                output_dir=Path(td),
                report_name="payments",
                sheet_name="Оплаты",
                columns=[ExportColumn("a", "Колонка A")],
                rows=[],
                now=datetime(2026, 5, 21, 8, 30),
            )
            self.assertTrue(out.exists())
            wb = load_workbook(out)
            ws = wb["Оплаты"]
            self.assertEqual(ws["A1"].value, "Колонка A")
            self.assertEqual(ws["A2"].value, "Данные отсутствуют")

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
    def test_export_with_data(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = export_rows_to_xlsx(
                output_dir=Path(td),
                report_name="cards",
                sheet_name="Карточки",
                columns=[ExportColumn("num", "Номер"), ExportColumn("fio", "ФИО")],
                rows=[{"num": "000001", "fio": "Иванов Иван"}],
                now=datetime(2026, 5, 21, 8, 31),
            )
            wb = load_workbook(out)
            ws = wb["Карточки"]
            self.assertEqual(ws["A1"].value, "Номер")
            self.assertEqual(ws["B1"].value, "ФИО")
            self.assertEqual(ws["A2"].value, "000001")
            self.assertEqual(ws["B2"].value, "Иванов Иван")


if __name__ == "__main__":
    unittest.main()
