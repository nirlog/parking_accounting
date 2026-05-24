from __future__ import annotations

from datetime import date, datetime
from importlib.util import find_spec
from pathlib import Path
import tempfile
import unittest

from parking_app.services.export_service import export_rows_to_xlsx
from parking_app.services.payments_export_service import (
    build_payments_export_rows,
    payments_export_columns,
)
from parking_app.services.payments_table_service import PaymentTableRow

OPENPYXL_AVAILABLE = find_spec("openpyxl") is not None
if OPENPYXL_AVAILABLE:
    from openpyxl import load_workbook


class PaymentsExportServiceTests(unittest.TestCase):
    def _row(self, *, status: str = "active", note: str = "Комментарий") -> PaymentTableRow:
        return PaymentTableRow(
            payment_id=1,
            payment_date=date(2026, 5, 21),
            period_from=date(2026, 5, 1),
            period_to=date(2026, 5, 31),
            amount_kopecks=800000,
            fio="Иванов Иван",
            state_number="А123АА178",
            place_number="147",
            receipt_number="483",
            fiscal_number="ФД85",
            accepted_by="Колобков",
            status=status,
            note=note,
        )

    def test_build_payments_export_rows_formats_dates_amount_and_status(self) -> None:
        rows = build_payments_export_rows([self._row(status="active")])
        self.assertEqual(rows[0]["payment_date"], "21.05.2026")
        self.assertEqual(rows[0]["amount"], "8 000.00")
        self.assertEqual(rows[0]["status"], "Активная")

    def test_build_payments_export_rows_formats_cancelled_status(self) -> None:
        rows = build_payments_export_rows([self._row(status="cancelled")])
        self.assertEqual(rows[0]["status"], "Отменена")

    def test_payments_export_columns_order(self) -> None:
        keys = [c.key for c in payments_export_columns()]
        self.assertEqual(
            keys,
            [
                "payment_date",
                "period_from",
                "period_to",
                "amount",
                "fio",
                "state_number",
                "place_number",
                "receipt_number",
                "fiscal_number",
                "accepted_by",
                "status",
                "note",
            ],
        )

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
    def test_export_service_still_sanitizes_formula_values(self) -> None:
        export_rows = build_payments_export_rows([self._row(note="=1+1")])
        with tempfile.TemporaryDirectory() as td:
            out = export_rows_to_xlsx(
                output_dir=Path(td),
                report_name="payments",
                sheet_name="Оплаты",
                columns=payments_export_columns(),
                rows=export_rows,
                now=datetime(2026, 5, 21, 8, 30),
            )
            wb = load_workbook(out)
            ws = wb["Оплаты"]
            self.assertEqual(ws["L2"].value, "'=1+1")


if __name__ == "__main__":
    unittest.main()
